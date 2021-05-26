import traceback
import datetime
import re,os.path
from tkinter import Tk, Menu, mainloop, filedialog, messagebox
import tkinter as tk
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Color
from openpyxl.styles import colors
from openpyxl.comments import Comment
import re

import openpyxl as xl

# create excel file from plain text file
def createExcel():
    # select text file
    filePath = filedialog.askopenfilename(initialdir="./", title="Select Input File",
                                      filetypes=(("Text Files", "*.txt"), ("all files", "*.*")))
    file=open(filePath)
    # read first line from inp file
    header=file.readline()
    # header identifier constant value 30
    headerIdentifier=header[:2]
    # if value is not 30 show error
    if headerIdentifier!="30":
        tk.messagebox.showerror(title="Invalid Header",message="invalide header value in text file")
    #extract orignmator code responder code file refernce number and no of records  from inp file
    orignatorCode=header[2:13]
    responderCode=header[13:24]
    fileReferenceNumber=header[24:34]
    totalRecords=header[34:40]
    # check if total Records are 000000
    test = re.search("^(?!0{6})",totalRecords)
    if not test:
        tk.messagebox.showerror(title="Invalid Data", message="invalide No of Records code in text file")
    # create excel file
    wb = xl.Workbook()
    ws = wb.active
    # Save Heder Information in Excel File
    ws.cell(1, 1).value = "Header Identifier"
    ws.cell(2, 1).value = "Orignator Code"
    ws.cell(3, 1).value = "Reponder Code"
    ws.cell(4, 1).value = "File Ref No"
    ws.cell(5, 1).value = "Total Records in File"

    ws.cell(1, 2).value=headerIdentifier
    ws.cell(2, 2).value = orignatorCode
    ws.cell(3, 2).value = responderCode
    ws.cell(4, 2).value = fileReferenceNumber
    ws.cell(5, 2).value = totalRecords

    # AT row 6 col 1 create header row
    r=6
    c=1
    ws.cell(r, c).value = "Account Valid Flag"
    c=c+1
    ws.cell(r, c).value = "Joint Account Flag"
    c = c + 1
    ws.cell(r, c).value = "Primary A/C Holder's PAN No"
    c = c + 1
    ws.cell(r, c).value = "Secondary A/C Holder's PAN No"
    c = c + 1
    ws.cell(r, c).value = "Primary Account Holder's Name"
    c = c + 1
    ws.cell(r, c).value = "Secondary Account Holder's Name"
    c = c + 1
    ws.cell(r, c).value = "Account type"
    c = c + 1
    ws.cell(r, c).value = "Record  Reference Number"
    c = c + 1
    ws.cell(r, c).value = "IFSC code of the bank branch at which customer account is maintained"
    c = c + 1
    ws.cell(r, c).value = "Destination Bank Account Number"
    c = c + 1
    ws.cell(r, c).value = "Filler 1"
    c = c + 1
    ws.cell(r, c).value = "Filler 2"
    c = c + 1
    ws.cell(r, c).value = "Filler 3"
    c = c + 1
    ws.cell(r, c).value = "Filler 4"
    c = c + 1
    ws.cell(r, c).value = "Filler 5"
    c = c + 1
    ws.cell(r, c).value = "Filler 6"
    c = c + 1
    ws.cell(r, c).value = "Filler 7"
    c = c + 1
    ws.cell(r, c).value = "Filler 8"
    c = c + 1
    ws.cell(r, c).value = "Filler 9"

    # at row 7 extract records from text file
    r=7
    c=1
    # accout valid flag validation
    avf = DataValidation(type="list", formula1='"00,01,02"', allow_blank=True)
    ws.add_data_validation(avf)
    avf.prompt="Please Select Account validation Flag"
    avf.promptTitle="Please Select from List"
    # joint account validation flags
    javf = DataValidation(type="list", formula1='"00,01"', allow_blank=True)
    ws.add_data_validation(javf)
    # account type validation flag
    at = DataValidation(type="list", formula1='"SB,CA,CC,OD,TD,LN,SG,CG,OT,NR,PP,NO"', allow_blank=True)
    ws.add_data_validation(at)

    atv = DataValidation(type="list", formula1='"00,01,51,52,53,54,55,60,62,65,68,69"', allow_blank=True)
    ws.add_data_validation(atv)
    # loop throgh inp file starting from 2nd line onwards
    for x in range(int(totalRecords)):
        line=file.readline()
        recordIdentifier=line[:2]
        recordRefNo=line[2:17]
        ifscCode=line[17:28]
        destinationBankAccountNo=line[28:63]
        filler1=line[67:87]
        filler2=line[87:137]
        filler3 = line[137:187]
        filler4 = line[187:237]
        filler5 = line[237:287]
        filler6 = line[287:337]
        filler7 = line[337:387]
        filler8 = line[387:437]
        filler9 = line[437:500]

        # first column in excel file account validation flag
        avf.add(ws.cell(r, c))
        ws.cell(r,c).number_format = '@'
        c = c + 1

        # second column in excel file joint account validation flag
        javf.add(ws.cell(r, c))
        ws.cell(r, c).number_format = '@'
        c = c + 1

        # Third column in excel file Primary A/C Holder's PAN No
        ws.cell(r, c).number_format = '@'
        c = c + 1

        # Fourth column in excel file Secondary A/C Holder's PAN No
        ws.cell(r, c).number_format = '@'
        c = c + 1

        # Fifth column in excel file Primary Account Holder's Name
        ws.cell(r, c).number_format = '@'
        c = c + 1

        # Sixth column in excel file Secondary Account Holder's Name
        ws.cell(r, c).number_format = '@'
        c = c + 1

        # Seventh column in excel file Account type
        at.add(ws.cell(r, c))
        ws.cell(r, c).number_format = '@'
        c = c + 1

        # Eighth column in excel file Record  Reference Number
        ws.cell(r, c).value = recordRefNo
        c = c + 1

        # Ninth column in excel file Record  IFSC code of the bank branch at which customer account is maintained
        ws.cell(r, c).value = ifscCode
        c = c + 1

        # Tenth column in excel file Record  Destination Bank Account Number
        ws.cell(r, c).value = destinationBankAccountNo
        c = c + 1

        # Eleventh column in excel file Record Filler 1
        ws.cell(r, c).value = filler1
        c = c + 1

        # Eleventh column in excel file Record Filler 2
        ws.cell(r, c).value = filler2
        c = c + 1

        # Eleventh column in excel file Record Filler 3
        ws.cell(r, c).value = filler3
        c = c + 1

        # Eleventh column in excel file Record Filler 4
        ws.cell(r, c).value = filler4
        c = c + 1

        # Eleventh column in excel file Record Filler 5
        ws.cell(r, c).value = filler5
        c = c + 1

        # Eleventh column in excel file Record Filler 6
        ws.cell(r, c).value = filler6
        c = c + 1

        # Eleventh column in excel file Record Filler 7
        ws.cell(r, c).value = filler7
        c = c + 1

        # Eleventh column in excel file Record Filler 8
        ws.cell(r, c).value = filler8
        c = c + 1

        # Eleventh column in excel file Record Filler 9
        ws.cell(r, c).value = filler9
        r=r+1
        c=1
    # save file at path from where input file is selected
    fileName = os.path.basename(filePath)
    ws.cell(3, 3).value = "FCBX68"
    ws.cell(3, 4).value = fileName[23:31]
    ws.cell(3, 5).value = fileName[32:37
                          ]
    wb.save(filePath+".xlsx")
    tk.messagebox.showinfo(title="Excel File", message="File Created sucessfuly at path from where input file was selected")



# create response file
def createResponse():
    filePath = filedialog.askopenfilename(initialdir="./", title="Select Input File",
                                          filetypes=(("Text Files", "*.xlsx"), ("all files", "*.*")))

    wb=xl.load_workbook(filePath)

    sheet=wb.active
    cell=sheet.cell(row=1,column=2)
    fileName="AV-"+ sheet.cell(row=2, column=2).value[:4] + "-" + sheet.cell(row=3, column=2).value[:4] + "-" + str(sheet.cell(row=3, column=3).value) + "-" + str(sheet.cell(row=3, column=4).value) + "-" +  str(sheet.cell(row=3, column=5).value) + "-RES.txt"
    f = open(fileName, "w")
    f.write(cell.value)
    cell = sheet.cell(row=2, column=2)
    txt=cell.value

    f.write(cell.value)
    cell = sheet.cell(row=3, column=2)
    f.write(cell.value)
    cell = sheet.cell(row=4, column=2)
    f.write(cell.value)
    cell = sheet.cell(row=5, column=2)
    f.write(cell.value)
    data = "                                                  "
    f.write(data)
    f.write(data)
    f.write(data)
    f.write(data)
    f.write(data)
    f.write(data)
    f.write(data)
    f.write(data)
    f.write(data)
    data = "          "
    f.write(data)
    r=7
    f.write("\n")

    for x in range(int(cell.value)):

        cell.value = "70"
        f.write(cell.value)
        cell = sheet.cell(row=r, column=8)
        spaces = " "
        if len(cell.value) < 15:
            spaces = spaces * (15 - len(cell.value))
            cell.value = cell.value + spaces
        f.write(cell.value)
        cell = sheet.cell(row=r, column=9)
        spaces = " "
        if len(cell.value) < 11:
            spaces = spaces * (11 - len(cell.value))
            cell.value = cell.value + spaces
        f.write(cell.value)
        cell = sheet.cell(row=r, column=10)
        spaces = " "
        if len(cell.value) < 35:
            spaces = spaces * (35 - len(cell.value))
            cell.value = cell.value + spaces
        f.write(cell.value)

        cell = sheet.cell(row=r, column=1)
        f.write(cell.value)

        cell = sheet.cell(row=r, column=2)
        if cell.value==None:
            cell.value="  "

        f.write(cell.value)
        cell = sheet.cell(row=r, column=3)
        if cell.value==None:
            cell.value="          "
        f.write(cell.value)

        cell = sheet.cell(row=r, column=4)
        if cell.value==None:
            cell.value="          "
        f.write(cell.value)
        cell = sheet.cell(row=r, column=5)
        if cell.value==None:
            cell.value="                                                  "
        spaces = " "
        if len(cell.value) < 50:
            spaces = spaces * (50 - len(cell.value))
            cell.value = cell.value + spaces

        f.write(cell.value)

        cell = sheet.cell(row=r, column=6)

        if cell.value == None:
            cell.value = "                                                  "
        spaces = " "
        if len(cell.value) < 50:
            spaces = spaces * (50 - len(cell.value))
            cell.value = cell.value + spaces
        f.write(cell.value)
        cell = sheet.cell(row=r, column=7)
        if cell.value == None:
            cell.value = "  "
        f.write(cell.value)

        cell = sheet.cell(row=r, column=11)
        if cell.value==None:
            cell.value="                                                  "
        spaces = " "
        if len(cell.value) < 50:
            spaces = spaces * (50 - len(cell.value))
            cell.value = cell.value + spaces
        f.write(cell.value)
        cell = sheet.cell(row=r, column=12)
        if cell.value==None:
            cell.value="                                                  "
        spaces = " "
        if len(cell.value) < 50:
            spaces = spaces * (50 - len(cell.value))
            cell.value = cell.value + spaces
        f.write(cell.value)
        cell = sheet.cell(row=r, column=13)
        if cell.value==None:
            cell.value="                                                  "
        spaces = " "
        if len(cell.value) < 50:
            spaces = spaces * (50 - len(cell.value))
            cell.value = cell.value + spaces
        f.write(cell.value)
        cell = sheet.cell(row=r, column=14)
        if cell.value==None:
            cell.value="                                                  "
        spaces = " "
        if len(cell.value) < 50:
            spaces = spaces * (50 - len(cell.value))
            cell.value = cell.value + spaces
        f.write(cell.value)
        cell = sheet.cell(row=r, column=15)
        if cell.value==None:
            cell.value="                                                  "
        spaces = " "
        if len(cell.value) < 50:
            spaces = spaces * (50 - len(cell.value))
            cell.value = cell.value + spaces
        f.write(cell.value)
        cell = sheet.cell(row=r, column=16)
        if cell.value == None:
            cell.value = "                                                  "
        spaces = " "
        if len(cell.value) < 50:
            spaces = spaces * (50 - len(cell.value))
            cell.value = cell.value + spaces
        f.write(cell.value)
        cell = sheet.cell(row=r, column=17)
        if cell.value==None or len(cell.value) > 11:
            cell.value="           "
        spaces = " "
        if len(cell.value) < 11:
            spaces = spaces * (11 - len(cell.value))
            cell.value = cell.value + spaces
        f.write(cell.value)

        r=r+1
        f.write("\n")

    f.close()
    tk.messagebox.showinfo(title="Response File", message="File Created sucessfuly")
main_window=Tk()
main_window.title="Utility for MIcrosoft Excel"
main_window.attributes("-fullscreen",True)



menu=Menu(main_window)
main_window.config(menu=menu)
splitMenu=Menu(menu)
menu.add_cascade(label="CDBT", menu=splitMenu)
splitMenu.add_command(label="Create Excel",command=createExcel)
splitMenu.add_command(label="Create Response",command=createResponse)
mainloop()